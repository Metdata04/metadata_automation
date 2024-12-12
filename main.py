import requests
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

# Configuration for a Akurana station
station_name = "FECT - Akurana"
mac_address = "80:7D:3A:7C:4F:08"  # MAC address of the station

# Directory for storing reports
output_dir = "station_reports"
os.makedirs(output_dir, exist_ok=True)

# API credentials
API_BASE_URL = "https://api.ambientweather.net/v1"  #API base URL
API_KEY = os.getenv("API_KEY", "35014481aea047c89787384ca605b666f4e569d67afb474a9071359c89aca8bd")  

def fetch_station_data(mac_address, year, month):
    """Fetch data for a station using its MAC address."""
    endpoint = f"{API_BASE_URL}/stations"
    params = {
        "macAddress": mac_address,
        "year": year,
        "month": month,
        "apikey": API_KEY
    }
    response = requests.get(endpoint, params=params)
    if response.status_code != 200:
        print(f"Error fetching data for MAC {mac_address}: {response.status_code}")
        return pd.DataFrame()
    try:
        data = response.json()
        return pd.DataFrame(data["records"])  
    except Exception as e:
        print(f"Error parsing data for MAC {mac_address}: {e}")
        return pd.DataFrame()

def append_to_excel(file_path, data):
    """Append data to an existing Excel file or create a new one."""
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            data.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    else:
        data.to_excel(file_path, index=False)

def process_station(station_name, mac_address):
    """Process data for a single station."""
    now = datetime.now()
    year = now.year
    month = now.month  # Numeric month for easier sorting
    print(f"Processing data for {station_name}...")
    station_data = fetch_station_data(mac_address, year, month)
    if station_data.empty:
        print(f"No data for {station_name}.")
        return
    file_path = os.path.join(output_dir, f"{station_name.replace(' ', '_')}_Metadata.xlsx")
    append_to_excel(file_path, station_data)
    print(f"Data for {station_name} saved at {file_path}.")

def main():
    """Main workflow."""
    print(f"Starting data processing for {station_name}...")
    process_station(station_name, mac_address)
    print("Processing completed.")

if __name__ == "__main__":
    main()
